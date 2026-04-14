from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.workbook import Workbook


@dataclass
class ExtractionRule:
    header_name: str
    extract_instruction: str
    enabled: bool = True


def _normalize(value: object) -> str:
    return str(value).strip() if value is not None else ""


def _default_instruction(header_name: str) -> str:
    normalized = header_name.strip()
    lower = normalized.lower()
    if any(keyword in normalized for keyword in ["日期", "时间"]) or "date" in lower:
        return f"从文档中提取{normalized}，统一输出为 YYYY-MM-DD。"
    if any(keyword in normalized for keyword in ["电话", "手机", "联系方式"]) or "phone" in lower:
        return f"从文档中提取{normalized}，仅输出号码本身。"
    if any(keyword in normalized for keyword in ["金额", "价税", "预算", "费用"]) or "amount" in lower:
        return f"从文档中提取{normalized}，保留金额和币种/单位。"
    if any(keyword in normalized for keyword in ["编号", "合同号", "单号"]) or "no" in lower or "code" in lower:
        return f"从文档中提取{normalized}，保持原始编号格式。"
    return f"从文档中提取{normalized}。"


def load_template_headers(path: Path) -> tuple[Workbook, list[str]]:
    workbook = load_workbook(path)
    worksheet = workbook.active
    first_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not first_row:
        raise ValueError("Excel file is empty.")
    headers = [_normalize(cell) for cell in first_row if _normalize(cell)]
    if not headers:
        raise ValueError("The first row of the Excel template must contain at least one header.")
    return workbook, headers


def infer_mappings(headers: list[str]) -> list[dict[str, object]]:
    return [
        {
            "header_name": header,
            "extract_instruction": _default_instruction(header),
            "enabled": True,
        }
        for header in headers
    ]


def write_results_to_template(
    workbook: Workbook,
    headers: list[str],
    rows: list[dict[str, object]],
) -> Workbook:
    sheet = workbook.active
    existing_headers = [_normalize(cell.value) for cell in sheet[1] if _normalize(cell.value)]
    final_headers = list(headers)
    if "document_name" not in existing_headers:
        sheet.insert_cols(1)
        sheet.cell(row=1, column=1, value="document_name")
        final_headers = ["document_name", *headers]
    else:
        final_headers = existing_headers

    if sheet.max_row > 1:
        sheet.delete_rows(2, sheet.max_row - 1)

    for row_index, row in enumerate(rows, start=2):
        values = row.get("values", {})
        for col_index, header in enumerate(final_headers, start=1):
            if header == "document_name":
                value = row.get("document_name", "")
            else:
                value = values.get(header, "")
            sheet.cell(row=row_index, column=col_index, value=value)

    return workbook
